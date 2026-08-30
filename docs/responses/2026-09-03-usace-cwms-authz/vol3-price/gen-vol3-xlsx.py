#!/usr/bin/env python3
"""Builds out/Vol-III-Price-FrasierDigital.xlsx from model.json — formula-driven, all 8 CLINs,
rates by labor category, hours per person per task (instr. 2.3.1; Q&A 5)."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
M = json.load(open(os.path.join(HERE, "model.json")))
OUT = os.path.join(HERE, "..", "out"); os.makedirs(OUT, exist_ok=True)
PATH = os.path.join(OUT, "Vol-III-Price-FrasierDigital.xlsx")
BOLD = Font(name="Times New Roman", bold=True, size=11); BASE = Font(name="Times New Roman", size=11)
H1 = Font(name="Times New Roman", bold=True, size=13); GREY = PatternFill("solid", fgColor="E8E8E8")
thin = Side(style="thin"); BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY, MONEY0 = '#,##0.00', '#,##0'
PEOPLE = ["Joseph", "Efrain", "Scott", "Zach", "Randy", "Andrew", "Seth"]
NAMES = {"Joseph": "J. Frasier", "Efrain": "E. Rocha", "Scott": "S. Carpenter", "Zach": "Z. Antosko", "Randy": "R. Chong", "Andrew": "A. Frasier", "Seth": "S. Chesky"}

def style(ws, row, cols, bold=False, fill=False):
    for c in cols:
        cell = ws.cell(row=row, column=c); cell.font = BOLD if bold else BASE
        if fill: cell.fill = GREY
        cell.border = BOX; cell.alignment = Alignment(vertical="top", wrap_text=True)

wb = Workbook()
# ---- Sheet: Rates ----
wr = wb.active; wr.title = "Labor Rates"
wr["A1"] = "Frasier Digital, LLC — Fully Burdened Labor Rates"; wr["A1"].font = H1
wr["A2"] = "Solicitation PANHEC-26-P-0000-026407 — CWMS Database Authorization Maintenance and Improvements"; wr["A2"].font = BASE
for i, h in enumerate(["Person", "Labor category", "Key personnel", "Fully burdened rate ($/hr)"], 1): wr.cell(row=4, column=i, value=h)
style(wr, 4, range(1, 5), bold=True, fill=True)
RATE_ROW = {}
r = 5
for p in PEOPLE:
    cat = M["cat"][p]
    wr.cell(row=r, column=1, value=NAMES[p]); wr.cell(row=r, column=2, value=M["label"][p])
    wr.cell(row=r, column=3, value="Yes" if cat != "AS" else "No")
    wr.cell(row=r, column=4, value=M["rate"][cat]).number_format = MONEY
    style(wr, r, range(1, 5)); RATE_ROW[p] = r; r += 1
for col, w in zip("ABCD", [16, 44, 14, 24]): wr.column_dimensions[col].width = w

# ---- Sheet: Hours by task ----
wh = wb.create_sheet("Hours by Task")
wh["A1"] = "Level of Effort — hours per person per task (identical to Volume I, Factor 1, Table A-2)"; wh["A1"].font = H1
hdr = ["CLIN", "Task", "Description", "Option"] + [NAMES[p] for p in PEOPLE] + ["Total hours", "Task price"]
for i, h in enumerate(hdr, 1): wh.cell(row=3, column=i, value=h)
style(wh, 3, range(1, len(hdr) + 1), bold=True, fill=True)
r = 4; first = r; PRICE_CELL = {}
for t in M["tasks"]:
    wh.cell(row=r, column=1, value=t["clin"]); wh.cell(row=r, column=2, value=t["task"])
    wh.cell(row=r, column=3, value=t["desc"]); wh.cell(row=r, column=4, value="Yes" if t["option"] else "No")
    terms = []
    for j, p in enumerate(PEOPLE):
        col = 5 + j; h = t["hours"].get(p, 0)
        wh.cell(row=r, column=col, value=h)
        cl = wh.cell(row=r, column=col).column_letter
        terms.append(f"{cl}{r}*'Labor Rates'!$D${RATE_ROW[p]}")
    tc = 5 + len(PEOPLE)
    wh.cell(row=r, column=tc, value=f"=SUM({wh.cell(row=r, column=5).column_letter}{r}:{wh.cell(row=r, column=tc-1).column_letter}{r})")
    wh.cell(row=r, column=tc + 1, value="=" + "+".join(terms)).number_format = MONEY0
    PRICE_CELL[t["clin"]] = f"'Hours by Task'!{wh.cell(row=r, column=tc+1).column_letter}{r}"
    style(wh, r, range(1, len(hdr) + 1)); r += 1
last = r - 1
wh.cell(row=r, column=3, value="Total — base and all options")
for j in range(len(PEOPLE) + 2):
    col = 5 + j; cl = wh.cell(row=r, column=col).column_letter
    wh.cell(row=r, column=col, value=f"=SUM({cl}{first}:{cl}{last})")
wh.cell(row=r, column=5 + len(PEOPLE) + 1).number_format = MONEY0
style(wh, r, range(1, len(hdr) + 1), bold=True, fill=True)
widths = [7, 9, 44, 8] + [12] * len(PEOPLE) + [12, 14]
for i, w in enumerate(widths, 1): wh.column_dimensions[wh.cell(row=3, column=i).column_letter].width = w

# ---- Sheet: CLIN Summary ----
ws = wb.create_sheet("CLIN Summary", 0)
ws["A1"] = "Frasier Digital, LLC — Price Proposal (Volume III, Factor 4)"; ws["A1"].font = H1
ws["A2"] = "Solicitation PANHEC-26-P-0000-026407 — CWMS Database Authorization Maintenance and Improvements"; ws["A2"].font = BASE
ws["A3"] = "Firm-Fixed-Price per task CLIN · Period of performance 360 days from award · Options per RFO 52.217-7"; ws["A3"].font = BASE
for i, h in enumerate(["CLIN", "Task", "Description", "Qty", "Unit", "Price"], 1): ws.cell(row=5, column=i, value=h)
style(ws, 5, range(1, 7), bold=True, fill=True)
r = 6; base_rows = []; opt_rows = []
for t in M["tasks"]:
    ws.cell(row=r, column=1, value=t["clin"]); ws.cell(row=r, column=2, value=t["task"]); ws.cell(row=r, column=3, value=t["desc"])
    ws.cell(row=r, column=4, value=1); ws.cell(row=r, column=5, value="Job")
    ws.cell(row=r, column=6, value="=" + PRICE_CELL[t["clin"]]).number_format = MONEY0
    (opt_rows if t["option"] else base_rows).append(r); style(ws, r, range(1, 7)); r += 1
ws.cell(row=r, column=3, value="Base tasks subtotal (CLINs 1001, 2001, 3001, 5001)")
ws.cell(row=r, column=6, value="=" + "+".join(f"F{x}" for x in base_rows)).number_format = MONEY0
style(ws, r, range(1, 7), bold=True, fill=True); rb = r; r += 1
ws.cell(row=r, column=3, value="Option tasks subtotal (CLINs 1002, 3002, 4001, 5002)")
ws.cell(row=r, column=6, value="=" + "+".join(f"F{x}" for x in opt_rows)).number_format = MONEY0
style(ws, r, range(1, 7), bold=True, fill=True); ro = r; r += 1
ws.cell(row=r, column=3, value="GRAND TOTAL — base and all options")
ws.cell(row=r, column=6, value=f"=F{rb}+F{ro}").number_format = MONEY0
style(ws, r, range(1, 7), bold=True, fill=True)
for col, w in zip("ABCDEF", [7, 9, 60, 6, 6, 14]): ws.column_dimensions[col].width = w
wb.save(PATH); print("wrote", PATH)
