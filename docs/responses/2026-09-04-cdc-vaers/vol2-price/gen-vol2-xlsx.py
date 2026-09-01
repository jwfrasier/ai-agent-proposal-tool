#!/usr/bin/env python3
"""Builds out/Vol-II-Price-Quote-FrasierDigital.xlsx — editable, formula-driven,
itemized per RFQ Vol II rules (labor cats/hours/rates per CLIN; ODCs separately priced)."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "..", "out")
os.makedirs(OUT, exist_ok=True)
PATH = os.path.join(OUT, "Vol-II-Price-Quote-FrasierDigital.xlsx")

BOLD = Font(name="Times New Roman", bold=True, size=11)
BASE = Font(name="Times New Roman", size=11)
H1 = Font(name="Times New Roman", bold=True, size=13)
GREY = PatternFill("solid", fgColor="E8E8E8")
thin = Side(style="thin")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '#,##0.00'
MONEY0 = '#,##0'

LINES = [
    ("Program Manager / Lead Architect", "Key Personnel", 546, 235.00),
    ("Security & Compliance Lead (SA&A / ATD / EPLC artifacts)", "Key Personnel (dual role)", 340, 235.00),
    ("Technical Lead — Healthcare IT", "Key Personnel", 390, 205.00),
    ("Integration & Data Engineer", "", 234, 175.00),
    ("Frontend Engineer / Test Automation", "", 780, 135.00),
    ("QA & Accessibility Engineer", "", 300, 135.00),
    ("Content & UX Research Specialist", "", 190, 105.00),
]

wb = Workbook()

def style_row(ws, row, cols, bold=False, fill=False, border=True):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD if bold else BASE
        if fill: cell.fill = GREY
        if border: cell.border = BOX
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# ---- Sheet 1: CLIN Summary ----
ws = wb.active; ws.title = "CLIN Summary"
ws["A1"] = "Frasier Digital, LLC — Price Quote"; ws["A1"].font = H1
ws["A2"] = "RFQ 75D301-26-Q-00146 — Modernized VAERS Reporting Application"; ws["A2"].font = BASE
ws["A3"] = "Firm-Fixed-Price · Period of Performance 9/28/2026–6/27/2027 (9 months) · Payment monthly in arrears"; ws["A3"].font = BASE
hdr = ["CLIN", "Description", "Type", "Price"]
for i, h in enumerate(hdr, 1): ws.cell(row=5, column=i, value=h)
style_row(ws, 5, range(1, 5), bold=True, fill=True)
rows = [
    ("0001", "Design, development, deployment, and compliance delivery of the Modernized VAERS Reporting Application per the PWS (severable services)", "FFP", "='CLIN 0001 Itemization'!E14"),
    ("0002", "Travel — onboarding (direct reimbursement per FTR, no fee)", "NTE", 31500),
    ("0003", "Travel — meetings (direct reimbursement per FTR, no fee)", "NTE", 31500),
]
r = 6
for row in rows:
    for i, v in enumerate(row, 1): ws.cell(row=r, column=i, value=v)
    ws.cell(row=r, column=4).number_format = MONEY0
    style_row(ws, r, range(1, 5)); r += 1
ws.cell(row=r, column=2, value="Total (CLIN 0001 + travel NTE ceilings)")
ws.cell(row=r, column=4, value=f"=SUM(D6:D{r-1})").number_format = MONEY0
style_row(ws, r, range(1, 5), bold=True, fill=True)
for col, w in zip("ABCD", [8, 80, 10, 14]): ws.column_dimensions[col].width = w

# ---- Sheet 2: CLIN 0001 Itemization ----
ws2 = wb.create_sheet("CLIN 0001 Itemization")
ws2["A1"] = "CLIN 0001 — Labor Itemization (Firm-Fixed-Price)"; ws2["A1"].font = H1
ws2["A2"] = "Fully burdened firm-fixed billing rates; hours reconcile with the staffing commitments in Volume I, Tab 3-1."; ws2["A2"].font = BASE
hdr = ["Labor Category", "Designation", "Hours", "Fully Burdened Rate ($/hr)", "Extended Price"]
for i, h in enumerate(hdr, 1): ws2.cell(row=4, column=i, value=h)
style_row(ws2, 4, range(1, 6), bold=True, fill=True)
r = 5
for cat, desig, hrs, rate in LINES:
    ws2.cell(row=r, column=1, value=cat)
    ws2.cell(row=r, column=2, value=desig)
    ws2.cell(row=r, column=3, value=hrs)
    ws2.cell(row=r, column=4, value=rate).number_format = MONEY
    ws2.cell(row=r, column=5, value=f"=C{r}*D{r}").number_format = MONEY0
    style_row(ws2, r, range(1, 6)); r += 1
ws2.cell(row=r, column=1, value="Subtotal")
ws2.cell(row=r, column=3, value=f"=SUM(C5:C{r-1})")
ws2.cell(row=r, column=5, value=f"=SUM(E5:E{r-1})").number_format = MONEY0
style_row(ws2, r, range(1, 6), bold=True, fill=True); r += 1
ws2.cell(row=r, column=1, value="Firm-fixed-price adjustment (rounding to CLIN price)")
ws2.cell(row=r, column=5, value=f"=E{r+1}-E{r-1}").number_format = MONEY0
style_row(ws2, r, range(1, 6)); r += 1
ws2.cell(row=r, column=1, value="CLIN 0001 FIRM FIXED PRICE")
ws2.cell(row=r, column=5, value=495000).number_format = MONEY0
style_row(ws2, r, range(1, 6), bold=True, fill=True)
for col, w in zip("ABCDE", [50, 24, 10, 22, 16]): ws2.column_dimensions[col].width = w

# ---- Sheet 3: ODCs / IT supplies-services ----
ws3 = wb.create_sheet("ODCs and IT Items")
ws3["A1"] = "Other Direct Costs and Separately Priced IT Supplies / Services"; ws3["A1"].font = H1
hdr = ["Item", "Basis", "Price"]
for i, h in enumerate(hdr, 1): ws3.cell(row=3, column=i, value=h)
style_row(ws3, 3, range(1, 4), bold=True, fill=True)
odc = [
    ("Software licenses, subscriptions, IT supplies", "None required — all custom code delivered open source per M-16-21; development on contractor equipment; laptops and PIV cards are Government-furnished", 0),
    ("Cloud hosting and AI services", "None required — CDC-managed Azure environment and CDC enterprise Azure OpenAI (EDAV) are Government-furnished", 0),
    ("Other direct costs (excluding travel CLINs 0002/0003)", "None", 0),
]
r = 4
for row in odc:
    for i, v in enumerate(row, 1): ws3.cell(row=r, column=i, value=v)
    ws3.cell(row=r, column=3).number_format = MONEY0
    style_row(ws3, r, range(1, 4)); r += 1
ws3.cell(row=r, column=1, value="Total ODCs")
ws3.cell(row=r, column=3, value=0).number_format = MONEY0
style_row(ws3, r, range(1, 4), bold=True, fill=True)
for col, w in zip("ABC", [45, 90, 12]): ws3.column_dimensions[col].width = w

wb.save(PATH)
print("wrote", PATH)
